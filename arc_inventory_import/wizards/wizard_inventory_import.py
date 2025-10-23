# -*- coding: utf-8 -*-
import logging
import base64
import pytz

from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

from odoo import models, fields, api, _
from odoo.tools import ustr, DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT

_logger = logging.getLogger(__name__)


class WizardInventoryImport(models.TransientModel):
    _name = 'wizard.inventory.import'
    _description = _('Asistente para importar productos del inventario')
    
    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        res_id = self.env.context.get('active_id')
        res_model = self.env.context.get('active_model')
        if res_id and res_model:
            res.update({
                'res_model': res_model,
                'res_id': res_id
            })
        return res

    def get_date_utc(self, date=False):
        if self.env.user.partner_id.tz:
            user_time_zone = pytz.timezone(self.env.user.partner_id.tz)
        else:
            user_time_zone = pytz.timezone('America/Lima')
        if date:
            current_date = datetime(date.year, date.month, date.day).date()
        else:
            current_date = datetime.now(user_time_zone).date()
        return current_date
    
    res_model = fields.Char("Modelo")
    res_id = fields.Integer("Id")
    
    name = fields.Char(string="Nombre del Inventario")
    file = fields.Binary(string="Archivo")
    date = fields.Date(string="Fecha del inventario", default=get_date_utc, required=True)
    location_id = fields.Many2one('stock.location', string="Ubicación")
    product_type = fields.Selection([
        ('name', 'Nombre'),
        ('barcode', 'Código de barra'),
        ('code', 'Código de producto'),
        ('minicode', 'Minicodigo')], string='Importar productos por', default='code')
    import_option = fields.Selection([('xls', 'Excel')], string='Tipo de archivo', default='xls')
    serial_lot = fields.Boolean(string="Crear número de serie si no existe", default=True)

    def show_success_msg(self, counter, skipped_line_no):
        dic_msg = "%s Registros sincronizados con éxito" % counter
        if skipped_line_no:
            dic_msg = dic_msg + "\nDetalle:"
            for k,v in skipped_line_no.items():
                dic_msg = dic_msg + "\nFila N° " + k + " " + v + " "
        context = dict(self._context or {})
        context['message'] = dic_msg
        view = self.env.ref("sh_message.sh_message_wizard")
        return {
            "name": "Importación de Productos",
            "view_mode": "form",
            "res_model": "sh.message.wizard",
            "type": "ir.actions.act_window",
            "views": [(view.id, "form")],
            "view_id": view.id,
            "target": "new",
            "context": context,
        }
    
    def read_xls(self):
        decoded_file = base64.decodebytes(self.file)
        excel_file = BytesIO(decoded_file)
        workbook = load_workbook(excel_file)

        sheet = workbook.active
        values_sheet = []
        
        for rowx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.value is None:
                    values.append("")
                elif cell.data_type == 'n':  # Número
                    is_float = cell.value % 1 != 0.0 if cell.value is not None else False
                    values.append(
                        str(cell.value)
                        if is_float
                        else str(int(cell.value))
                    )
                elif cell.is_date:  # Fechas
                    dt = cell.value
                    if isinstance(dt, datetime.datetime):
                        is_datetime = dt.time() != datetime.time(0, 0, 0)
                        values.append(
                            dt.strftime(DEFAULT_SERVER_DATETIME_FORMAT)
                            if is_datetime
                            else dt.strftime(DEFAULT_SERVER_DATE_FORMAT)
                        )
                    else:
                        values.append(str(cell.value))
                elif cell.data_type == 'b':  # Booleano
                    values.append(u'True' if cell.value else u'False')
                elif cell.data_type == 'e':  # Error en la celda
                    raise ValueError(
                        _("Valor de celda no válido en la fila %(row)s, columna %(col)s: %(cell_value)s") % {
                            'row': rowx,
                            'col': colx,
                            'cell_value': cell.value
                        }
                    )
                else:  # Texto u otro tipo de datos
                    values.append(cell.value if cell.value is not None else "")
            values_sheet.append(values)
        return values_sheet
       
    def import_stock_inventory_line(self, values, inventory):
        location_ids = inventory.location_ids.mapped("id")
        company = inventory.company_id
        
        product_id = False
        field_search_value = False
        product_product = self.env['product.product']
        field_search = ""
        if self.product_type == 'minicode':
            product_id = product_product.search([('minicode', '=', values.get('minicode'))], limit=1)
            field_search = "Minicodigo"
            field_search_value = values.get('minicode')
        elif self.product_type == 'code':
            product_id = product_product.search([('default_code', '=', values.get('default_code'))], limit=1)
            field_search = "Código"
            field_search_value = values.get('default_code')
        elif self.product_type == 'name':
            product_id = product_product.search([('name', 'ilike', values.get('name'))], limit=1)
            field_search = "Nombre"
            field_search_value = values.get('name')
        elif self.product_type == 'barcode':
            product_id = product_product.search([('barcode', '=', values.get('barcode'))], limit=1)
            field_search = "Código de barra"
            field_search_value = values.get('barcode')
                
        vals = {}
        message = ""
        
        if product_id:
            if self.serial_lot and product_id.tracking == 'serial':
                serial_number = values.get('lot_id')
                if serial_number not in (None, ""):
                    lot_id = self.env['stock.lot'].search([
                        ('name', '=', serial_number.strip()),
                        ('product_id', '=', product_id.id),
                        ('company_id', '=', company.id),
                        ('location_id', '=', location_ids[0])
                    ])
                    vals_lot = {}

                    if not lot_id:
                        vals_lot = {
                            'name': serial_number,
                            'product_id': product_id.id,
                            'company_id': company.id,
                            'location_id': location_ids[0],
                        }
                        try:
                            lot_id = self.env['stock.lot'].create(vals_lot)
                            vals.update({'lot_id': lot_id.id})
                        except Exception as e:
                            return "Error al crear la serie %s del producto %s. Error %s" % (serial_number, product_id.name, ustr(e))
                    else:
                        vals.update({'lot_id': lot_id.id})

            quantity = 1
            if values.get('product_qty') not in (None, ""):
                quantity = int(values.get('product_qty').strip())
            
            vals.update({
                'product_id': product_id.id,
                'location_id': location_ids[0],
                'quantity': quantity,
                'reserved_quantity': 0.0,
                'company_id': company.id,
                'inventory_date': self.date,
            })

            try:
                if product_id.tracking == 'serial':
                    domain = [
                        ('product_id', '=', product_id.id),
                        ('location_id', '=', location_ids[0]),
                        ('lot_id', '=', vals.get('lot_id'))
                    ]
                else:
                    domain = [
                        ('product_id', '=', product_id.id),
                        ('location_id', '=', location_ids[0]),
                    ]
                if not self.env['stock.quant'].search(domain):
                    self.env['stock.quant'].create(vals)
                else:
                    self.env['stock.quant'].write({
                        'quantity': quantity,
                    })
                return True
            except Exception as e:
                return "Error al crear un item del inventario - Producto %s. Error %s" % (product_id.name, ustr(e))
        else:
            message = "Producto no encontrado: %s - %s" % (field_search, field_search_value)
        return message

    def action_import(self):
        _logger.info("========== action_import ==========")
        inventory_id = self.env[self.res_model].browse(self.res_id)
                
        counter = 1
        message = ""
        skipped_line_no = {}
        if self.import_option == 'xls':
            rows = self.read_xls()
            values = {}
            skip_header = True
            
            for row in rows:
                if skip_header:
                    skip_header = False
                    counter += 1
                    continue
                                
                values.update({
                    'name': row[0],
                    'product_qty': row[1],
                    'lot_id': row[2],
                    'default_code': row[3],
                    'minicode': row[4],
                    'barcode': row[5],
                })

                result = self.import_stock_inventory_line(values, inventory_id)
                if not result:
                    skipped_line_no[counter] = result
                counter = counter + 1
            if counter > 1:
                completed_records = (counter - len(skipped_line_no)) - 2
                message = self.show_success_msg(completed_records, skipped_line_no)
        return message
