# -*- coding: utf-8 -*-

import base64
import csv
import io
import logging

from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from odoo.tools import ustr, DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT

_logger = logging.getLogger(__name__)


class WizardInventoryVariantsImport(models.TransientModel):
    _name = 'wizard.inventory.variants.import'
    _description = 'Asistente para Importar Inventario'

    file_data = fields.Binary(
        string='Archivo',
        required=True,
        help="Archivo CSV con los datos de inventario. Columnas esperadas: Descripcion, Atributos, Valor de atributos, Cantidad"
    )
    location_id = fields.Many2one(
        'stock.location',
        string='Ubicación de existencias',
        required=True,
        domain="[('usage', '=', 'internal')]",
        help="Ubicación donde se actualizará el stock."
    )
    
    def show_success_msg(self, counter, skipped_line_no):
        dic_msg = "%s Registros sincronizados con éxito" % counter
        if skipped_line_no:
            dic_msg = dic_msg + "\nDetalle:"
            for k, v in skipped_line_no.items():
                dic_msg = dic_msg + "\nFila N° " + k + " " + v + " "
        context = dict(self._context or {})
        context['message'] = dic_msg
        view = self.env.ref("sh_message.sh_message_wizard")
        return {
            "name": "Inventario de variantes",
            "view_mode": "form",
            "res_model": "sh.message.wizard",
            "type": "ir.actions.act_window",
            "views": [(view.id, "form")],
            "view_id": view.id,
            "target": "new",
            "context": context,
        }
    
    def read_xls(self):
        decoded_file = base64.decodebytes(self.file_data)
        excel_file = BytesIO(decoded_file)
        workbook = load_workbook(excel_file)

        sheet = workbook.active
        values_sheet = []

        for rowx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.value is None:
                    values.append("")
                elif cell.data_type == 'n':  # Número
                    is_float = cell.value % 1 != 0.0 if cell.value is not None else False
                    values.append(
                        str(cell.value)
                        if is_float
                        else str(int(cell.value))
                    )
                elif cell.is_date:  # Fechas
                    dt = cell.value
                    if isinstance(dt, datetime.datetime):
                        is_datetime = dt.time() != datetime.time(0, 0, 0)
                        values.append(
                            dt.strftime(DEFAULT_SERVER_DATETIME_FORMAT)
                            if is_datetime
                            else dt.strftime(DEFAULT_SERVER_DATE_FORMAT)
                        )
                    else:
                        values.append(str(cell.value))
                elif cell.data_type == 'b':  # Booleano
                    values.append(u'True' if cell.value else u'False')
                elif cell.data_type == 'e':  # Error en la celda
                    raise ValueError(
                        _("Valor de celda no válido en la fila %(row)s, columna %(col)s: %(cell_value)s") % {
                            'row': rowx,
                            'col': colx,
                            'cell_value': cell.value
                        }
                    )
                else:  # Texto u otro tipo de datos
                    values.append(cell.value if cell.value is not None else "")
            values_sheet.append(values)
        return values_sheet

    def _find_product_variant(self, product_name, attribute_name, attribute_value_name):
        """
        Encuentra una variante de producto (product.product) específica basada en
        el nombre de la plantilla, el nombre del atributo y el nombre del valor del atributo.
        """
        # 1. Buscar la plantilla de producto por nombre.
        product_template = self.env['product.template'].search([('name', '=', product_name)], limit=1)
        if not product_template:
            return None

        # 2. Buscar las variantes que pertenecen a esta plantilla.
        domain = [('product_tmpl_id', '=', product_template.id)]
        product_variants = self.env['product.product'].search(domain)
        
        # 3. Filtrar las variantes para encontrar la que coincide con el atributo/valor.
        for variant in product_variants:
            # product_template_attribute_value_ids contiene la combinación de valores de atributo para esa variante.
            # Ejemplo: [('Color', 'Rojo'), ('Talla', 'M')]
            variant_attribute_values = variant.product_template_attribute_value_ids
            
            # Buscamos si alguno de los valores de atributo de la variante coincide con el que buscamos.
            match = variant_attribute_values.filtered(
                lambda val: val.attribute_id.name == attribute_name and val.product_attribute_value_id.name == attribute_value_name
            )
            
            # Si encontramos una coincidencia y la variante solo tiene ese atributo (o los que correspondan), es la correcta.
            # Para el caso de un solo atributo, si hay 'match', es suficiente.
            if match:
                # En un caso más complejo con múltiples atributos por producto, se necesitaría verificar que
                # todos los atributos del archivo coincidan para una misma variante.
                # Para el archivo de ejemplo, esta lógica es suficiente.
                return variant

        return None

    def action_import_inventory(self):
        """
        Procesa el archivo Excel subido para actualizar las cantidades de inventario
        de las variantes de producto en la ubicación seleccionada.
        """
        self.ensure_one()

        if not self.file_data:
            raise UserError(_("Por favor, suba un archivo para procesar."))

        try:
            skip_header = True
            counter = 1
            skipped_line_no = {}
            values = self.read_xls()
            errors = []
            
            for row in values:
                if skip_header:
                    skip_header = False
                    counter = counter + 1
                    continue
                
                if row[0] in (None, ""):
                    skipped_line_no[str(counter)] = " - Descripción esta vacio. "
                    counter = counter + 1
                else:
                    product_name = str(row[0]).strip()
                    attribute_name = str(row[6]).strip()
                    attribute_value = str(row[7]).strip()
                    
                    try:
                        quantity = float(row[9].strip())
                    except (ValueError, IndexError):
                        errors.append(f"Fila {counter}: Cantidad '{row[9]}' no válida para '{product_name}'.")
                        continue

                    if not all([product_name, attribute_name, attribute_value]):
                        errors.append(f"Fila {counter}: Faltan datos clave (Descripción, Atributo o Valor).")
                        continue

                    # --- Lógica para encontrar la variante de producto ---
                    product_variant = self._find_product_variant(product_name, attribute_name, attribute_value)

                    if not product_variant:
                        errors.append(f"Fila {counter}: No se encontró la variante para '{product_name}' con {attribute_name}='{attribute_value}'.")
                        continue
                    
                    # --- Actualizar inventario usando stock.quant ---
                    # Este método actualiza la cantidad o crea un nuevo quant si no existe.
                    self.env['stock.quant'].with_context(inventory_mode=True).create({
                        'product_id': product_variant.id,
                        'location_id': self.location_id.id,
                        'quantity': quantity,
                    })
                    _logger.info(f"Inventario actualizado para {product_variant.display_name}: {quantity} en {self.location_id.display_name}")

            if errors:
                error_message = "\n".join(errors)
                raise UserError(_("El proceso de importación finalizó con los siguientes errores:\n\n%s") % error_message)

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Importación Exitosa'),
                    'message': _('El inventario ha sido actualizado correctamente.'),
                    'type': 'success',
                    'sticky': False,
                }
            }

        except Exception as e:
            raise ValidationError(_("Lo sentimos, el excel no coincide con el formato \n" + ustr(e)))
