# -*- coding: utf-8 -*-

import base64
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

from odoo import models, fields, _
from odoo.exceptions import ValidationError
from odoo.tools import ustr, DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT

import logging
_logger = logging.getLogger(__name__)


class ProductImport(models.TransientModel):
    _name = "wizard.product.import"
    _description = "Importar productos con o sin series"

    company_id = fields.Many2one('res.company', string="Empresa", default=lambda self: self.env.company, required=True)
    file = fields.Binary(string="Archivo")
    serial_lot = fields.Boolean(string="Crear número de serie si no existe")
    product_type = fields.Selection([
        ('name', 'Nombre'),
        ('barcode', 'Código de barra'),
        ('code', 'Código de producto'),
        ('minicode', 'Minicodigo')], string='Importar productos por', default='code')
    import_option = fields.Selection([('xls', 'XLS')], string='Tipo de archivo', default='xls')
    import_action = fields.Selection([
        ('sync', 'Sincronizar'),
        ('reportproduct', 'Lista de Productos')
    ], string='Acción', required=True, default='sync')
    update = fields.Boolean(string="Actualizar campos específicos de productos")
    # fields sync
    field_name = fields.Boolean(string="Nombre")
    field_cost = fields.Boolean(string="Costo")
    field_price = fields.Boolean(string="Precios")
    field_category = fields.Boolean(string="Categorias")
    field_model = fields.Boolean(string="Modelo")
    field_minicode = fields.Boolean(string="Minicodigo")
    field_tecnology = fields.Boolean(string="Tecnología")
    field_default_code = fields.Boolean(string="Código")
    field_description_sale = fields.Boolean(string="Descripción de venta")
    field_tracking = fields.Boolean(string="Trazabilidad de series")

    def show_success_msg(self, counter, skipped_line_no):
        dic_msg = "%s Registros sincronizados con éxito" % counter
        if skipped_line_no:
            dic_msg = dic_msg + "\nDetalle:"
            for k, v in skipped_line_no.items():
                dic_msg = dic_msg + "\nFila N° " + k + " " + v + " "
        context = dict(self._context or {})
        context['message'] = dic_msg
        view = self.env.ref("sh_message.sh_message_wizard")
        return {
            "name": "Sincronización de productos",
            "view_mode": "form",
            "res_model": "sh.message.wizard",
            "type": "ir.actions.act_window",
            "views": [(view.id, "form")],
            "view_id": view.id,
            "target": "new",
            "context": context,
        }

    def action_sync(self):
        self.ensure_one()
        if not self.file:
            raise ValidationError("Para sincronizar debe subir un archivo Excel.")
        if self.update:
            if not self.field_name and not self.field_cost and not self.field_price and \
                not self.field_category and not self.field_model and not self.field_default_code and \
                    self.field_minicode:
                raise ValidationError("Debe seleccionar al menos un campo para actualizar.")

        if self.import_action == 'sync':
            counter, skipped_line_no = self.sync_products()
            if counter > 1:
                completed_records = (counter - len(skipped_line_no)) - 2
                res = self.show_success_msg(completed_records, skipped_line_no)
                return res

    def action_export(self):
        self.ensure_one()
        if self.import_action == 'reportproduct':
            return {
                'type': 'ir.actions.act_url',
                'url': '/report/products?company_id=' + str(self.company_id.id) + '',
                'target': 'new'
            }

    def read_xls(self):
        decoded_file = base64.decodebytes(self.file)
        excel_file = BytesIO(decoded_file)
        workbook = load_workbook(excel_file)

        sheet = workbook.active
        values_sheet = []

        for rowx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.value is None:
                    values.append("")
                elif cell.data_type == 'n':  # Número
                    is_float = cell.value % 1 != 0.0 if cell.value is not None else False
                    values.append(
                        str(cell.value)
                        if is_float
                        else str(int(cell.value))
                    )
                elif cell.is_date:  # Fechas
                    dt = cell.value
                    if isinstance(dt, datetime.datetime):
                        is_datetime = dt.time() != datetime.time(0, 0, 0)
                        values.append(
                            dt.strftime(DEFAULT_SERVER_DATETIME_FORMAT)
                            if is_datetime
                            else dt.strftime(DEFAULT_SERVER_DATE_FORMAT)
                        )
                    else:
                        values.append(str(cell.value))
                elif cell.data_type == 'b':  # Booleano
                    values.append(u'True' if cell.value else u'False')
                elif cell.data_type == 'e':  # Error en la celda
                    raise ValueError(
                        _("Valor de celda no válido en la fila %(row)s, columna %(col)s: %(cell_value)s") % {
                            'row': rowx,
                            'col': colx,
                            'cell_value': cell.value
                        }
                    )
                else:  # Texto u otro tipo de datos
                    values.append(cell.value if cell.value is not None else "")
            values_sheet.append(values)
        return values_sheet

    def create_categ_id(self, name, public=False):
        if public:
            categ_id = self.env['product.public.category'].search([('name', '=', name)])
            if categ_id:
                return categ_id
            else:
                categ_id = self.env['product.public.category'].create({'name': name})
                return categ_id
        else:
            categ_id = self.env['product.category'].search([('name', '=', name)])
            if categ_id:
                return categ_id
            else:
                categ_id = self.env['product.category'].create({'name': name})
                return categ_id

    def create_subcateg_id(self, name, parent_id, public=False):
        if public:
            categ_id = self.env['product.public.category'].search([('name', '=', name), ('parent_id', '=', parent_id.id)], limit=1)
            if categ_id:
                return categ_id
            else:
                categ_id = self.env['product.public.category'].create({'name': name, 'parent_id': parent_id.id})
                return categ_id
        else:
            categ_id = self.env['product.category'].search([('name', '=', name), ('parent_id', '=', parent_id.id)], limit=1)
            if categ_id:
                return categ_id
            else:
                categ_id = self.env['product.category'].create({'name': name, 'parent_id': parent_id.id})
                return categ_id

    def create_product(self, values):
        product_name = ""
        if values.get('product') in (None, ""):
            raise ValidationError("Se debe ingresar el nombre del producto en el archivo Excel.")
        else:
            product_name = values.get('product').strip().upper()

        minicode = values.get('minicode')
        product_id = self.env['product.template'].search([('minicode', '=', minicode)])
        if not product_id:
            vals = {}
            if values.get('category') not in (None, ""):
                categ_id = self.create_categ_id(values.get('category').strip())
                if values.get('subcategory').strip() not in (None, ""):
                    subcateg_id = self.create_subcateg_id(values.get('subcategory').strip(), categ_id)
                    vals.update({
                        'categ_id': subcateg_id.id
                    })
            list_price = 0.0
            if values.get('list_price') not in (None, ""):
                list_price = values.get('list_price')
                vals.update({
                    'list_price': list_price
                })
            if values.get('standard_price') not in (None, ""):
                vals.update({
                    'standard_price': values.get('standard_price'),
                })
            if values.get('model') not in (None, ""):
                vals.update({
                    'model': values.get('model').strip(),
                })
            if values.get('tecnology') not in (None, ""):
                vals.update({
                    'tecnology': values.get('tecnology').strip(),
                })
            if values.get('default_code') not in (None, ""):
                vals.update({
                    'default_code': values.get('default_code').strip().upper(),
                })
            if values.get('description_sale') not in (None, ""):
                vals.update({
                    'description_sale': values.get('description_sale').strip(),
                })

            trazabilidad = 'none'
            if values.get('lot') not in (None, ""):
                if int(values.get('lot')) == 1:
                    trazabilidad = 'serial'

            vals.update({
                'detailed_type': 'product',
                'name': product_name,
                'tracking': trazabilidad,
                'minicode': minicode,
                'list_price': list_price,
                'company_id': self.company_id.id
            })

            product_id = self.env['product.template'].create(vals)
            return product_id

    def update_product(self, values, product_id):
        vals = {}

        # Handle category updates
        if self.field_category and self._is_valid_string(values.get('category')):
            categ_id = self.create_categ_id(values.get('category'))

            if self._is_valid_string(values.get('subcategory')):
                subcateg_id = self.create_subcateg_id(
                    values.get('subcategory').strip(),
                    categ_id
                )
                vals['categ_id'] = subcateg_id.id

        # Handle price updates
        if (self.field_price and values.get('list_price') not in (None, "")):
            vals['list_price'] = values.get('list_price')

        # Handle cost updates
        if (self.field_cost and values.get('standard_price') not in (None, "")):
            vals['standard_price'] = values.get('standard_price')

        # Handle model updates
        if (self.field_model and self._is_valid_string(values.get('model'))):
            vals['model'] = values.get('model').strip()

        # Handle technology updates
        if self.field_tecnology:
            vals['tecnology'] = values.get('tecnology')

        # Handle name updates
        if (self.field_name and self._is_valid_string(values.get('product'))):
            vals['name'] = values.get('product').strip().upper()

        # Handle default code updates
        if (self.field_default_code and self._is_valid_string(values.get('default_code'))):
            vals['default_code'] = values.get('default_code').strip().upper()

        # Handle description sale updates
        if (self.field_description_sale and self._is_valid_string(values.get('description_sale'))):
            vals['description_sale'] = values.get('description_sale').strip()

        # Handle tracking updates
        if self.field_tracking:
            tracking_value = ('serial' if values.get('lot') not in (None, "") else 'none')
            vals['tracking'] = tracking_value

        # Handle minicode updates
        if (self.field_minicode and values.get('minicode') not in (None, "", 0)):
            vals['minicode'] = values.get('minicode')

        # Set product type
        vals['type'] = 'product'

        # Update product
        product_id.write(vals)
        return product_id

    def _is_valid_string(self, value):
        return value is not None and str(value).strip() != ""

    def sync_products(self):
        _logger.info("========== sync_products ==========")
        counter = 1
        skipped_line_no = {}
        if self.import_option == 'xls':
            rows = self.read_xls()
            values = {}
            skip_header = True

            try:
                for row in rows:
                    if skip_header:
                        skip_header = False
                        counter += 1
                        continue

                    field_search = "minicode"
                    field_search_value = ""

                    if row[2] not in (None, ""):
                        values = {
                            'product': row[0],
                            'default_code': row[1],
                            'minicode': row[2],
                            'lot': row[3],
                            'standard_price': row[4],
                            'list_price': row[5],
                            'description_sale': row[6],
                            'category': row[7],
                            'subcategory': row[8],
                            'tecnology': row[9],
                            'brand': row[10],
                            'public': row[11],
                            'model': row[12],
                            'warranty': row[13],
                            'availability': row[14],
                        }

                        product_obj = self.env['product.product']
                        field_search = ""
                        if self.product_type == 'minicode':
                            field_search = "Minicodigo"
                            field_search_value = values.get('minicode')
                            product_id = product_obj.search([('minicode', '=', values.get('minicode'))], limit=1)
                            if not product_id:
                                if values.get('name') not in (None, ""):
                                    product_id = product_obj.search([('name', '=', values.get('name'))], limit=1)
                                    if product_id:
                                        product_id.write({'minicode': values.get('minicode')})
                        elif self.product_type == 'code':
                            field_search = "Código"
                            field_search_value = values.get('default_code')
                            product_id = product_obj.search([('default_code', '=', values.get('default_code'))], limit=1)
                        elif self.product_type == 'name':
                            field_search = "Nombre"
                            field_search_value = values.get('name')
                            product_id = product_obj.search([('name', '=', values.get('name'))], limit=1)
                        elif self.product_type == 'barcode':
                            field_search = "Código de barra"
                            field_search_value = values.get('barcode')
                            product_id = product_obj.search([('barcode', '=', values.get('barcode'))], limit=1)

                        if product_id:
                            self.update_product(values, product_id)
                        else:
                            self.create_product(values)
                    else:
                        skipped_line_no[str(counter)] = " - %s: %s del producto no encontrado" % (field_search, field_search_value)

                    counter += 1
            except Exception as e:
                skipped_line_no[str(counter)] = " - Error: %s" % ustr(e)
                raise ValidationError("Lo sentimos, su archivo excel no coincide con el formato \n" + ustr(e))

        return counter, skipped_line_no
