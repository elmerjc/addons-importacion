# -*- coding: utf-8 -*-

import base64
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

from odoo import models, fields, _
from odoo.exceptions import ValidationError
from odoo.tools import ustr, DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT

import logging
_logger = logging.getLogger(__name__)


class ProductVariantImport(models.TransientModel):
    _name = "wizard.product.variant.import"
    _description = "Importar productos y variantes desde Excel"

    file = fields.Binary(string="Archivo", required=True)
    
    def show_success_msg(self, counter, skipped_line_no):
        dic_msg = "%s Registros sincronizados con éxito" % counter
        if skipped_line_no:
            dic_msg = dic_msg + "\nDetalle:"
            for k, v in skipped_line_no.items():
                dic_msg = dic_msg + "\nFila N° " + k + " " + v + " "
        context = dict(self._context or {})
        context['message'] = dic_msg
        view = self.env.ref("sh_message.sh_message_wizard")
        return {
            "name": "Sincronización de productos",
            "view_mode": "form",
            "res_model": "sh.message.wizard",
            "type": "ir.actions.act_window",
            "views": [(view.id, "form")],
            "view_id": view.id,
            "target": "new",
            "context": context,
        }
         
    def read_xls(self):
        decoded_file = base64.decodebytes(self.file)
        excel_file = BytesIO(decoded_file)
        workbook = load_workbook(excel_file)

        sheet = workbook.active
        values_sheet = []

        for rowx, row in enumerate(sheet.iter_rows(values_only=False), 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.value is None:
                    values.append("")
                elif cell.data_type == 'n':  # Número
                    is_float = cell.value % 1 != 0.0 if cell.value is not None else False
                    values.append(
                        str(cell.value)
                        if is_float
                        else str(int(cell.value))
                    )
                elif cell.is_date:  # Fechas
                    dt = cell.value
                    if isinstance(dt, datetime.datetime):
                        is_datetime = dt.time() != datetime.time(0, 0, 0)
                        values.append(
                            dt.strftime(DEFAULT_SERVER_DATETIME_FORMAT)
                            if is_datetime
                            else dt.strftime(DEFAULT_SERVER_DATE_FORMAT)
                        )
                    else:
                        values.append(str(cell.value))
                elif cell.data_type == 'b':  # Booleano
                    values.append(u'True' if cell.value else u'False')
                elif cell.data_type == 'e':  # Error en la celda
                    raise ValueError(
                        _("Valor de celda no válido en la fila %(row)s, columna %(col)s: %(cell_value)s") % {
                            'row': rowx,
                            'col': colx,
                            'cell_value': cell.value
                        }
                    )
                else:  # Texto u otro tipo de datos
                    values.append(cell.value if cell.value is not None else "")
            values_sheet.append(values)
        return values_sheet

    def action_import(self):
        product_tmpl_obj = self.env['product.template']

        if self and self.file:
            counter = 1
            skipped_line_no = {}

            try:
                values = self.read_xls()
                skip_header = True
                running_tmpl = None
                created_product_tmpl = False
                has_variant = False

                for row in values:
                    try:
                        if skip_header:
                            skip_header = False
                            counter = counter + 1
                            continue

                        if row[0] not in (None, ""):
                            var_vals = {}
                            
                            # Product Template Start
                            if row[0] != running_tmpl:
                                running_tmpl = row[0]
                                
                                tmpl_vals = {
                                    'name': row[0],
                                    'sale_ok': True,
                                    'purchase_ok': True,
                                    'detailed_type': 'product',
                                    'type': 'product',
                                    'invoice_policy': 'order',
                                    'company_id': self.env.company.id
                                }
                                
                                if row[1] not in (None, ""):
                                    tmpl_vals.update({'id_articulo': row[1]})
                                
                                if row[2] not in (None, ""):
                                    tmpl_vals.update({'minicode': row[2]})
                                
                                if row[3] not in (None, ""):
                                    tmpl_vals.update({'default_code': str(row[3])})

                                if row[4] not in (None, ""):
                                    tmpl_vals.update({'list_price': row[4]})
                                
                                if row[5] not in (None, ""):
                                    tmpl_vals.update({'standard_price': row[5]})
                                
                                if row[6].strip() in (None, "") or row[7].strip() in (None, ""):
                                    has_variant = False
                                else:
                                    has_variant = True
                                
                                if row[8] not in (None, ""):
                                    tmpl_vals.update({'barcode': row[8]})
                            
                                # ===================================================================
                                # Step 1: Create Product Template
                                search_product = product_tmpl_obj.search([('name', '=', running_tmpl)], limit=1)
                                if search_product:
                                    # Write product Template Field.
                                    search_product.write(tmpl_vals)
                                    created_product_tmpl = search_product
                                else:
                                    created_product_tmpl = product_tmpl_obj.create(tmpl_vals)
                            
                            # Variant Values
                            if created_product_tmpl and has_variant:
                                # Product Variants
                                pro_attr_line_obj = self.env['product.template.attribute.line']
                                pro_attr_value_obj = self.env['product.attribute.value']
                                pro_attr_obj = self.env['product.attribute']
                                
                                if row[6].strip() not in (None, "") and row[7].strip() not in (None, ""):
                                    attr_ids_list = []
                                    for attr in row[6].split(','):
                                        attr = attr.strip()
                                        if attr != '':
                                            search_attr_name = False
                                            search_attr_name = pro_attr_obj.search([('name', '=', attr)], limit=1)
                                            if not search_attr_name:
                                                search_attr_name = pro_attr_obj.create({'name': attr})
                                            attr_ids_list.append(search_attr_name.id)
                                    
                                    attr_value_list = []
                                    attr_value_price_dic = {}
                                    for attr_value in row[7].split(','):
                                        attr_value = attr_value.strip()
                                        splited_attr_value_price_list = []
                                    
                                        # Product Attribute Price
                                        if '@' in attr_value:
                                            splited_attr_value_price_list = attr_value.split('@')
                                            attr_value_price_dic.update({
                                                splited_attr_value_price_list[0]: splited_attr_value_price_list[1]
                                            })
                                        else:
                                            splited_attr_value_price_list = [attr_value]

                                        if splited_attr_value_price_list[0] != '':
                                            attr_value_list.append(splited_attr_value_price_list[0])
                                                                            
                                    attr_value_ids_list = []
                                    if len(attr_ids_list) == len(attr_value_list):
                                        i = 0
                                        while i < len(attr_ids_list):
                                            search_attr_value = False
                                            search_attr_value = pro_attr_value_obj.search([
                                                ('name', '=', attr_value_list[i]),
                                                ('attribute_id', '=', attr_ids_list[i])
                                            ], limit=1)
                                                                                        
                                            if not search_attr_value:
                                                search_attr_value = pro_attr_value_obj.create({
                                                    'name' : attr_value_list[i],
                                                    'attribute_id' : attr_ids_list[i]
                                                })
                                            
                                            attr_value_ids_list.append(search_attr_value.id)
                                            i += 1
                                    else:
                                        skipped_line_no[str(counter)] = " - Número de atributos y su valor no es igual. "
                                        counter = counter + 1
                                        continue

                                    if attr_value_ids_list and attr_ids_list:
                                        i = 0
                                        while i < len(attr_ids_list):
                                            search_attr_line = pro_attr_line_obj.search([
                                                ('attribute_id', '=', attr_ids_list[i]),
                                                ('product_tmpl_id', '=', created_product_tmpl.id)
                                            ], limit=1)

                                            if search_attr_line:
                                                past_values_list = []
                                                past_values_list = search_attr_line.value_ids.ids
                                                past_values_list.append(attr_value_ids_list[i])
                                                search_attr_line.write({'value_ids': [(6, 0, past_values_list)]})
                                            else:
                                                pro_attr_line_obj.create({
                                                    'attribute_id' : attr_ids_list[i],
                                                    'value_ids': [(6, 0, [attr_value_ids_list[i]])],
                                                    'product_tmpl_id' : created_product_tmpl.id,
                                                })
                                            i += 1
                                    
                                    created_product_tmpl._create_variant_ids()
                                    
                                    if created_product_tmpl.product_variant_ids:
                                        product_var_obj = self.env['product.product']
                                        domain = []
                                        domain.append(('product_tmpl_id', '=', created_product_tmpl.id))
                                        
                                        for attr_value_id in attr_value_ids_list:
                                            domain.append(
                                                ('product_template_attribute_value_ids.product_attribute_value_id.id', '=', attr_value_id)
                                            )
                                        
                                        product_varient = product_var_obj.search(domain, limit=1)
                                        
                                        if not product_varient:
                                            if attr_value_ids_list and attr_ids_list:
                                                i = 0
                                                while i < len(attr_ids_list):
                                                    search_attr_line = pro_attr_line_obj.search([
                                                        ('attribute_id', '=', attr_ids_list[i]),
                                                        ('product_tmpl_id', '=', created_product_tmpl.id),
                                                    ], limit=1)
                                                    if search_attr_line:
                                                        past_values_list = []
                                                        past_values_list = search_attr_line.value_ids.ids
                                                        past_values_list.append(attr_value_ids_list[i])
                                                        search_attr_line.write({
                                                            'value_ids': [(6, 0, past_values_list)]
                                                        })
                                                    else:
                                                        pro_attr_line_obj.create({
                                                            'attribute_id' : attr_ids_list[i],
                                                            'value_ids': [(6, 0, [attr_value_ids_list[i]])],
                                                            'product_tmpl_id' : created_product_tmpl.id,
                                                        })
                                                    i += 1
                                            created_product_tmpl._create_variant_ids()
                                            product_varient = product_var_obj.search(domain, limit=1)

                                        if not product_varient:
                                            skipped_line_no[str(counter)] = " - Variantes de producto no encontradas."
                                            counter = counter + 1
                                            continue
                                        
                                        if row[1] not in (None, ""):
                                            var_vals.update({'id_articulo': row[1]})
                                        
                                        if row[2] not in (None, ""):
                                            var_vals.update({'minicode': row[2]})

                                        if row[3] not in (None, ""):
                                            var_vals.update({'default_code': str(row[3])})

                                        product_varient.write(var_vals)

                            counter = counter + 1
                        else:
                            skipped_line_no[str(counter)] = " - Descripción esta vacio. "
                            counter = counter + 1

                    except Exception as e:
                        skipped_line_no[str(counter)] = " - Valor no es valido. " + ustr(e)
                        counter = counter + 1
                        continue
                        
            except Exception as e:
                raise ValidationError(_("Lo sentimos, el excel no coincide con el formato \n" + ustr(e)))

            if counter > 1:
                completed_records = (counter - len(skipped_line_no)) - 2
                res = self.show_success_msg(completed_records, skipped_line_no)
                return res
